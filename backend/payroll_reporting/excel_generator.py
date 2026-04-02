"""
Excel/CSV Generator for Payroll Reports
Generates structured data export for payroll software
"""
import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import EmployeePayrollData, GlobalPayrollSummary


MONTHS_FR = {
    1: "Janvier", 2: "Fevrier", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Aout",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Decembre"
}


class PayrollExcelGenerator:
    """Generateur Excel pour les donnees de paie"""
    
    def __init__(self, hotel_name: str):
        self.hotel_name = hotel_name
        
        # Styles
        self.header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=10)
        self.title_font = Font(bold=True, size=14, color="1E293B")
        self.subtitle_font = Font(italic=True, size=10, color="64748B")
        self.total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        self.total_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
    
    def generate_excel(self, summary: GlobalPayrollSummary) -> bytes:
        """Genere le fichier Excel complet"""
        wb = Workbook()
        
        # Feuille 1: Donnees de paie detaillees
        ws_data = wb.active
        ws_data.title = "Variables de paie"
        self._create_payroll_sheet(ws_data, summary)
        
        # Feuille 2: Resume par service
        ws_summary = wb.create_sheet("Resume par service")
        self._create_summary_sheet(ws_summary, summary)
        
        # Feuille 3: Absences
        ws_absences = wb.create_sheet("Absences")
        self._create_absences_sheet(ws_absences, summary)
        
        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_payroll_sheet(self, ws, summary: GlobalPayrollSummary):
        """Cree la feuille principale des variables de paie"""
        month_name = MONTHS_FR.get(summary.month, str(summary.month))
        
        # Titre
        ws.merge_cells('A1:N1')
        ws['A1'] = f"VARIABLES DE PAIE - {self.hotel_name}"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:N2')
        ws['A2'] = f"{month_name} {summary.year}"
        ws['A2'].font = self.subtitle_font
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # En-tetes (ligne 4)
        headers = [
            "ID Salarie",
            "Nom",
            "Prenom", 
            "Mois",
            "H. Normales",
            "H. Sup 25%",
            "H. Sup 50%",
            "H. Nuit",
            "J. Travailles",
            "CP (jours)",
            "Maladie (jours)",
            "Abs. Non Just.",
            "Autres Abs.",
            "Total Heures"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Donnees
        for row_idx, emp in enumerate(summary.employees, 5):
            data = [
                emp.employee_id[:8] + "...",  # ID tronque
                emp.last_name,
                emp.first_name,
                f"{summary.month:02d}/{summary.year}",
                round(emp.normal_hours, 2),
                round(emp.overtime_25_hours, 2),
                round(emp.overtime_50_hours, 2),
                round(emp.night_hours, 2),
                emp.worked_days,
                round(emp.cp_days, 1),
                round(emp.sick_days, 1),
                round(emp.unjustified_absences, 1),
                round(emp.other_absences, 1),
                round(emp.total_hours_to_pay, 2)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center' if col > 3 else 'left')
        
        # Ligne totaux
        total_row = 5 + len(summary.employees)
        totals = [
            "TOTAUX",
            "",
            "",
            "",
            round(summary.total_normal_hours, 2),
            round(summary.total_overtime_25, 2),
            round(summary.total_overtime_50, 2),
            round(summary.total_night_hours, 2),
            summary.total_worked_days,
            round(summary.total_cp_days, 1),
            round(summary.total_sick_days, 1),
            round(summary.total_unjustified, 1),
            round(summary.total_other_absences, 1),
            round(summary.total_worked_hours, 2)
        ]
        
        for col, value in enumerate(totals, 1):
            cell = ws.cell(row=total_row, column=col, value=value)
            cell.fill = self.total_fill
            cell.font = self.total_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center' if col > 3 else 'left')
        
        # Ajuster largeurs colonnes
        column_widths = [12, 15, 12, 10, 12, 12, 12, 10, 12, 12, 14, 14, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Figer les volets
        ws.freeze_panes = 'A5'
    
    def _create_summary_sheet(self, ws, summary: GlobalPayrollSummary):
        """Cree la feuille resume par service"""
        month_name = MONTHS_FR.get(summary.month, str(summary.month))
        
        # Titre
        ws.merge_cells('A1:E1')
        ws['A1'] = f"RESUME PAR SERVICE - {month_name} {summary.year}"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # En-tetes
        headers = ["Service", "Effectif", "Heures Totales", "Heures Sup.", "% du Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Donnees par departement
        dept_names = {
            "front_office": "Reception",
            "housekeeping": "Hebergement", 
            "food_beverage": "Restauration",
            "maintenance": "Maintenance",
            "administration": "Direction",
            "kitchen": "Cuisine"
        }
        
        row = 4
        for dept, data in summary.by_department.items():
            dept_name = dept_names.get(dept, dept)
            hours = data.get('hours', 0)
            pct = (hours / summary.total_worked_hours * 100) if summary.total_worked_hours > 0 else 0
            
            values = [
                dept_name,
                int(data.get('count', 0)),
                round(hours, 2),
                round(data.get('overtime', 0), 2),
                f"{pct:.1f}%"
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
            row += 1
        
        # Ajuster largeurs
        for col, width in enumerate([18, 10, 14, 14, 12], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_absences_sheet(self, ws, summary: GlobalPayrollSummary):
        """Cree la feuille des absences"""
        month_name = MONTHS_FR.get(summary.month, str(summary.month))
        
        # Titre
        ws.merge_cells('A1:G1')
        ws['A1'] = f"DETAIL ABSENCES - {month_name} {summary.year}"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # En-tetes
        headers = ["Nom", "Prenom", "Service", "CP", "Maladie", "Non Just.", "Autres"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Donnees
        dept_names = {
            "front_office": "Reception",
            "housekeeping": "Hebergement",
            "food_beverage": "Restauration",
            "maintenance": "Maintenance",
            "administration": "Direction",
            "kitchen": "Cuisine"
        }
        
        row = 4
        for emp in summary.employees:
            if emp.total_absences > 0:  # Seulement ceux avec absences
                values = [
                    emp.last_name,
                    emp.first_name,
                    dept_names.get(emp.department, emp.department),
                    round(emp.cp_days, 1),
                    round(emp.sick_days, 1),
                    round(emp.unjustified_absences, 1),
                    round(emp.other_absences, 1)
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center' if col > 2 else 'left')
                row += 1
        
        # Totaux
        totals = [
            "TOTAUX",
            "",
            "",
            round(summary.total_cp_days, 1),
            round(summary.total_sick_days, 1),
            round(summary.total_unjustified, 1),
            round(summary.total_other_absences, 1)
        ]
        
        for col, value in enumerate(totals, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = self.total_fill
            cell.font = self.total_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center' if col > 2 else 'left')
        
        # Ajuster largeurs
        for col, width in enumerate([15, 12, 14, 8, 10, 10, 10], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def generate_csv(self, summary: GlobalPayrollSummary) -> str:
        """Genere le fichier CSV (alternative simple)"""
        lines = []
        
        # En-tete
        headers = [
            "ID_Salarie", "Nom", "Prenom", "Mois", "Annee",
            "H_Normales", "H_Sup_25", "H_Sup_50", "H_Nuit",
            "J_Travailles", "CP_Jours", "Maladie_Jours",
            "Abs_Non_Justifiees", "Autres_Absences", "Total_Heures"
        ]
        lines.append(";".join(headers))
        
        # Donnees
        for emp in summary.employees:
            row = [
                emp.employee_id,
                emp.last_name,
                emp.first_name,
                str(summary.month),
                str(summary.year),
                f"{emp.normal_hours:.2f}",
                f"{emp.overtime_25_hours:.2f}",
                f"{emp.overtime_50_hours:.2f}",
                f"{emp.night_hours:.2f}",
                str(emp.worked_days),
                f"{emp.cp_days:.1f}",
                f"{emp.sick_days:.1f}",
                f"{emp.unjustified_absences:.1f}",
                f"{emp.other_absences:.1f}",
                f"{emp.total_hours_to_pay:.2f}"
            ]
            lines.append(";".join(row))
        
        return "\n".join(lines)
