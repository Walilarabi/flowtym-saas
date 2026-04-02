"""
Flowtym Configuration Module - Excel Import Service

Provides functionality to import rooms from Excel files.
"""
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..models.rooms import (
    RoomBulkImport, RoomStatus, ViewType, BathroomType
)
from .validation import ValidationService

logger = logging.getLogger(__name__)


class ExcelImportError(Exception):
    """Custom exception for Excel import errors"""
    def __init__(self, message: str, errors: List[Dict] = None):
        super().__init__(message)
        self.errors = errors or []


class ExcelImportService:
    """Service for importing configuration data from Excel files"""
    
    # Expected columns for room import
    ROOM_COLUMNS = {
        "room_number": ["numero", "numero_chambre", "room_number", "chambre", "n°"],
        "room_type_code": ["type", "code_type", "room_type", "type_chambre", "code"],
        "floor": ["etage", "floor", "niveau"],
        "view": ["vue", "view"],
        "bathroom": ["salle_de_bain", "bathroom", "sdb"],
        "equipment": ["equipements", "equipment", "amenities", "équipements"],
        "is_accessible": ["pmr", "accessible", "handicap"],
        "notes": ["notes", "remarques", "comments"]
    }
    
    # Value mappings
    VIEW_MAPPING = {
        "ville": ViewType.CITY,
        "city": ViewType.CITY,
        "mer": ViewType.SEA,
        "sea": ViewType.SEA,
        "jardin": ViewType.GARDEN,
        "garden": ViewType.GARDEN,
        "piscine": ViewType.POOL,
        "pool": ViewType.POOL,
        "montagne": ViewType.MOUNTAIN,
        "mountain": ViewType.MOUNTAIN,
        "cour": ViewType.COURTYARD,
        "courtyard": ViewType.COURTYARD,
        "rue": ViewType.STREET,
        "street": ViewType.STREET,
        "parc": ViewType.PARK,
        "park": ViewType.PARK,
        "": ViewType.NONE,
        "aucune": ViewType.NONE,
        "none": ViewType.NONE,
    }
    
    BATHROOM_MAPPING = {
        "douche": BathroomType.SHOWER,
        "shower": BathroomType.SHOWER,
        "baignoire": BathroomType.BATHTUB,
        "bathtub": BathroomType.BATHTUB,
        "bath": BathroomType.BATHTUB,
        "les_deux": BathroomType.BOTH,
        "both": BathroomType.BOTH,
        "jacuzzi": BathroomType.JACUZZI,
    }
    
    BOOLEAN_TRUE = ["oui", "yes", "1", "true", "x", "o", "vrai"]
    BOOLEAN_FALSE = ["non", "no", "0", "false", "", "faux"]
    
    @classmethod
    def parse_rooms_excel(
        cls,
        file_content: bytes,
        tenant_id: str,
        existing_room_types: Dict[str, str]  # code -> id mapping
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Parse an Excel file containing room data.
        
        Args:
            file_content: Raw bytes of the Excel file
            tenant_id: The hotel's tenant ID
            existing_room_types: Mapping of room type codes to their IDs
            
        Returns:
            Tuple of (valid_rooms, errors)
            - valid_rooms: List of room dictionaries ready for insertion
            - errors: List of error dictionaries with row numbers and messages
        """
        try:
            wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            
            if ws is None:
                raise ExcelImportError("Le fichier Excel ne contient pas de feuille de données")
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                value = str(cell.value or "").lower().strip()
                # Remove accents and special chars for matching
                value = value.replace("é", "e").replace("è", "e").replace("ê", "e")
                value = value.replace("°", "").replace(" ", "_")
                headers.append(value)
            
            # Map headers to expected columns
            column_mapping = cls._map_columns(headers)
            
            if "room_number" not in column_mapping:
                raise ExcelImportError(
                    "Colonne 'numero' ou 'room_number' non trouvée",
                    [{"row": 1, "error": "La colonne du numéro de chambre est requise"}]
                )
            
            if "room_type_code" not in column_mapping:
                raise ExcelImportError(
                    "Colonne 'type' ou 'room_type' non trouvée",
                    [{"row": 1, "error": "La colonne du type de chambre est requise"}]
                )
            
            # Parse data rows
            valid_rooms = []
            errors = []
            seen_room_numbers = set()
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                
                # Skip completely empty rows
                if all(v is None or v == "" for v in row_data.values()):
                    continue
                
                try:
                    room = cls._parse_room_row(
                        row_data, 
                        column_mapping, 
                        tenant_id,
                        existing_room_types
                    )
                    
                    # Check for duplicate room numbers
                    if room["room_number"] in seen_room_numbers:
                        errors.append({
                            "row": row_idx,
                            "field": "room_number",
                            "value": room["room_number"],
                            "error": f"Numéro de chambre en double: {room['room_number']}"
                        })
                        continue
                    
                    seen_room_numbers.add(room["room_number"])
                    valid_rooms.append(room)
                    
                except ValueError as e:
                    errors.append({
                        "row": row_idx,
                        "error": str(e),
                        "data": {k: str(v)[:50] for k, v in row_data.items() if v}
                    })
            
            wb.close()
            
            return valid_rooms, errors
            
        except Exception as e:
            logger.error(f"Excel parsing error: {str(e)}")
            if isinstance(e, ExcelImportError):
                raise
            raise ExcelImportError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")
    
    @classmethod
    def _map_columns(cls, headers: List[str]) -> Dict[str, int]:
        """Map Excel column headers to expected field names"""
        mapping = {}
        
        for field, aliases in cls.ROOM_COLUMNS.items():
            for idx, header in enumerate(headers):
                if header in aliases or any(alias in header for alias in aliases):
                    mapping[field] = idx
                    break
        
        return mapping
    
    @classmethod
    def _parse_room_row(
        cls,
        row_data: Dict[str, Any],
        column_mapping: Dict[str, int],
        tenant_id: str,
        existing_room_types: Dict[str, str]
    ) -> Dict[str, Any]:
        """Parse a single row of room data"""
        
        def get_value(field: str, default=None):
            if field not in column_mapping:
                return default
            idx = column_mapping[field]
            keys = list(row_data.keys())
            if idx < len(keys):
                return row_data.get(keys[idx], default)
            return default
        
        # Room number (required)
        room_number = str(get_value("room_number", "")).strip()
        if not room_number:
            raise ValueError("Le numéro de chambre est requis")
        
        is_valid, error = ValidationService.validate_room_number(room_number)
        if not is_valid:
            raise ValueError(error)
        
        # Room type code (required)
        room_type_code = str(get_value("room_type_code", "")).strip().upper()
        if not room_type_code:
            raise ValueError("Le type de chambre est requis")
        
        # Validate room type exists
        if room_type_code not in existing_room_types:
            raise ValueError(f"Type de chambre inconnu: {room_type_code}. Types disponibles: {', '.join(existing_room_types.keys())}")
        
        room_type_id = existing_room_types[room_type_code]
        
        # Floor (optional, default 1)
        floor_raw = get_value("floor", 1)
        try:
            floor = int(float(floor_raw)) if floor_raw else 1
        except (ValueError, TypeError):
            floor = 1
        
        is_valid, error = ValidationService.validate_floor(floor)
        if not is_valid:
            raise ValueError(error)
        
        # View (optional)
        view_raw = str(get_value("view", "")).lower().strip()
        view = cls.VIEW_MAPPING.get(view_raw, ViewType.NONE)
        
        # Bathroom (optional)
        bathroom_raw = str(get_value("bathroom", "")).lower().strip()
        bathroom = cls.BATHROOM_MAPPING.get(bathroom_raw, BathroomType.SHOWER)
        
        # Equipment (optional, comma-separated)
        equipment_raw = get_value("equipment", "")
        equipment = []
        if equipment_raw:
            equipment = [e.strip() for e in str(equipment_raw).split(",") if e.strip()]
        
        # Is accessible (optional)
        accessible_raw = str(get_value("is_accessible", "")).lower().strip()
        is_accessible = accessible_raw in cls.BOOLEAN_TRUE
        
        # Notes (optional)
        notes = get_value("notes", "")
        if notes:
            notes = str(notes).strip()[:500]  # Limit length
        
        return {
            "room_number": room_number,
            "room_type_id": room_type_id,
            "room_type_code": room_type_code,
            "floor": floor,
            "view": view.value if hasattr(view, 'value') else view,
            "bathroom": bathroom.value if hasattr(bathroom, 'value') else bathroom,
            "equipment": equipment,
            "is_accessible": is_accessible,
            "notes": notes or None,
            "tenant_id": tenant_id,
            "status": RoomStatus.AVAILABLE.value
        }
    
    @classmethod
    def generate_template(cls) -> bytes:
        """
        Generate an Excel template for room import.
        Returns the file content as bytes.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Chambres"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            ("Numéro", 12),
            ("Type", 15),
            ("Étage", 10),
            ("Vue", 15),
            ("Salle de Bain", 18),
            ("Équipements", 40),
            ("PMR", 8),
            ("Notes", 30)
        ]
        
        for col, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Example data rows
        example_data = [
            ("101", "STD", 1, "Ville", "Douche", "wifi, tv, minibar", "Non", ""),
            ("102", "STD", 1, "Jardin", "Douche", "wifi, tv", "Non", "Vue sur le parc"),
            ("201", "SUP", 2, "Mer", "Baignoire", "wifi, tv, minibar, coffre", "Non", ""),
            ("202", "DLX", 2, "Mer", "Les deux", "wifi, tv, minibar, coffre, climatisation", "Oui", "Chambre PMR"),
            ("301", "STE", 3, "Ville", "Jacuzzi", "wifi, tv, minibar, coffre, climatisation, room service", "Non", "Suite Présidentielle"),
        ]
        
        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in [3, 7]:  # Étage, PMR
                    cell.alignment = Alignment(horizontal="center")
        
        # Instructions sheet
        ws_info = wb.create_sheet(title="Instructions")
        ws_info["A1"] = "Instructions d'importation des chambres"
        ws_info["A1"].font = Font(bold=True, size=14)
        
        instructions = [
            "",
            "Colonnes obligatoires:",
            "  - Numéro: Le numéro de la chambre (ex: 101, 201A, P-01)",
            "  - Type: Le code du type de chambre (doit exister dans la configuration)",
            "",
            "Colonnes optionnelles:",
            "  - Étage: Numéro de l'étage (défaut: 1)",
            "  - Vue: Ville, Mer, Jardin, Piscine, Montagne, Cour, Rue, Parc",
            "  - Salle de Bain: Douche, Baignoire, Les deux, Jacuzzi",
            "  - Équipements: Liste séparée par des virgules",
            "  - PMR: Oui/Non pour l'accessibilité",
            "  - Notes: Remarques libres",
            "",
            "Types de chambres courants:",
            "  - STD: Standard",
            "  - SUP: Supérieure",
            "  - DLX: Deluxe",
            "  - STE: Suite",
            "  - JST: Junior Suite",
            "",
            "Assurez-vous que les types de chambres sont créés",
            "dans la configuration avant l'import."
        ]
        
        for idx, text in enumerate(instructions, start=2):
            ws_info.cell(row=idx, column=1, value=text)
        
        ws_info.column_dimensions["A"].width = 60
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @classmethod
    def validate_import_preview(
        cls,
        valid_rooms: List[Dict],
        errors: List[Dict],
        existing_rooms: List[str]  # List of existing room numbers
    ) -> Dict[str, Any]:
        """
        Generate a preview/summary of the import.
        
        Args:
            valid_rooms: List of successfully parsed rooms
            errors: List of parsing errors
            existing_rooms: List of room numbers that already exist
            
        Returns:
            Import preview summary
        """
        # Check for conflicts with existing rooms
        conflicts = []
        new_rooms = []
        
        for room in valid_rooms:
            if room["room_number"] in existing_rooms:
                conflicts.append({
                    "room_number": room["room_number"],
                    "action": "update"  # Will update existing room
                })
            else:
                new_rooms.append(room)
        
        # Summary by type
        by_type = {}
        for room in valid_rooms:
            type_code = room.get("room_type_code", "UNKNOWN")
            by_type[type_code] = by_type.get(type_code, 0) + 1
        
        # Summary by floor
        by_floor = {}
        for room in valid_rooms:
            floor = room.get("floor", 1)
            by_floor[floor] = by_floor.get(floor, 0) + 1
        
        return {
            "total_rows": len(valid_rooms) + len(errors),
            "valid_count": len(valid_rooms),
            "error_count": len(errors),
            "new_rooms_count": len(new_rooms),
            "update_count": len(conflicts),
            "errors": errors[:20],  # Limit errors shown
            "by_type": by_type,
            "by_floor": dict(sorted(by_floor.items())),
            "conflicts": conflicts[:10],
            "can_proceed": len(valid_rooms) > 0
        }
